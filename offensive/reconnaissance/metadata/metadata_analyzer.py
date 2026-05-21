"""
Metadata Analyzer - تحليل البيانات الوصفية من الملفات

يستخرج:
- PDF metadata (المؤلف، البرنامج، التاريخ، العناوين)
- Office documents metadata (Word, Excel, PowerPoint)
- Image EXIF data (الكاميرا، GPS، التاريخ)
- أسماء المستخدمين المخفية
- المسارات الداخلية
- إصدارات البرمجيات المستخدمة
"""

import asyncio
import re
import io
import os
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from datetime import datetime
from urllib.parse import urljoin, urlparse

import httpx
import logging

logger = logging.getLogger(__name__)


@dataclass
class MetadataInfo:
    """معلومات البيانات الوصفية"""
    file_url: str
    file_type: str  # pdf, docx, xlsx, pptx, image, other
    file_size: int = 0
    
    # PDF metadata
    pdf_author: str = ""
    pdf_creator: str = ""
    pdf_producer: str = ""
    pdf_title: str = ""
    pdf_subject: str = ""
    pdf_created: str = ""
    pdf_modified: str = ""
    pdf_pages: int = 0
    
    # Office metadata
    office_author: str = ""
    office_last_modified_by: str = ""
    office_created: str = ""
    office_modified: str = ""
    office_company: str = ""
    office_manager: str = ""
    office_application: str = ""
    
    # Image EXIF
    image_make: str = ""
    image_model: str = ""
    image_software: str = ""
    image_datetime: str = ""
    image_gps_lat: str = ""
    image_gps_lon: str = ""
    image_gps_alt: str = ""
    
    # Extracted data
    usernames: List[str] = field(default_factory=list)
    email_addresses: List[str] = field(default_factory=list)
    internal_paths: List[str] = field(default_factory=list)
    software_versions: List[str] = field(default_factory=list)
    urls: List[str] = field(default_factory=list)
    
    # Raw text preview
    raw_text: str = ""
    findings: List[str] = field(default_factory=list)


@dataclass
class MetadataResult:
    """نتائج تحليل البيانات الوصفية"""
    target: str
    files_analyzed: int = 0
    metadata_found: List[MetadataInfo] = field(default_factory=list)
    total_usernames: int = 0
    total_emails: int = 0
    total_paths: int = 0
    total_versions: int = 0
    findings: List[Dict] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)


class MetadataAnalyzer:
    """
    محلل البيانات الوصفية
    
    يفحص الملفات المرفوعة علناً لاستخراج معلومات مخفية
    """
    
    # أنماط الملفات
    FILE_PATTERNS = [
        # PDFs
        r'\.pdf$',
        # Office documents
        r'\.docx?$', r'\.xlsx?$', r'\.pptx?$', r'\.od[tsp]$',
        # Images
        r'\.(?:jpg|jpeg|png|gif|tiff?|bmp)$',
        # Other
        r'\.(?:csv|txt|rtf|xml|json|yaml|yml)$',
    ]
    
    # أنماط استخراج البيانات
    USERNAME_PATTERNS = [
        r'(?:Author|Creator|LastModifiedBy|LastAuthor)[\s:]+([^\r\n]+)',
        r'/Users/([^/\s]+)',
        r'/home/([^/\s]+)',
        r'C:\\Users\\([^\\\s]+)',
        r'username[\s:=]+([^\s,;]+)',
    ]
    
    PATH_PATTERNS = [
        r'(?:/Users/[^\s,;]+)',
        r'(?:/home/[^\s,;]+)',
        r'(?:C:\\[^\s,;]+)',
        r'(?:/var/www/[^\s,;]+)',
        r'(?:/opt/[^\s,;]+)',
        r'(?:/etc/[^\s,;]+)',
    ]
    
    VERSION_PATTERNS = [
        r'(?:Microsoft\s+(?:Word|Excel|PowerPoint|Office)\s+(\d+[\d.]*))',
        r'(?:Adobe\s+(?:Acrobat|Photoshop|Illustrator)\s+(\d+[\d.]*))',
        r'(?:PDF\s*(?:Creator|Producer):\s*([^\r\n]+))',
        r'(?:Version:\s*([^\r\n]+))',
    ]
    
    def __init__(self):
        self._results: Dict[str, MetadataResult] = {}
        self._client = None
    
    async def _get_client(self) -> httpx.AsyncClient:
        if not self._client:
            self._client = httpx.AsyncClient(
                timeout=15, follow_redirects=True, verify=False,
                headers={"User-Agent": "Mozilla/5.0 (compatible; HunterMind/1.0)"}
            )
        return self._client
    
    async def analyze(self, url: str, html_content: str = "") -> MetadataResult:
        """
        تحليل البيانات الوصفية من الموقع
        
        Args:
            url: رابط الموقع
            html_content: HTML الصفحة (اختياري)
        
        Returns:
            MetadataResult مع البيانات المستخرجة
        """
        parsed = urlparse(url)
        base = f"{parsed.scheme}://{parsed.netloc}"
        
        print(f"  📋 Metadata Analysis: {base}")
        
        result = MetadataResult(target=base)
        client = await self._get_client()
        
        # 1. الحصول على HTML إذا لم يقدم
        if not html_content:
            try:
                response = await client.get(base)
                html_content = response.text
            except Exception as e:
                result.errors.append(str(e))
                return result
        
        # 2. استخراج روابط الملفات من HTML
        file_urls = self._extract_file_urls(html_content, base)
        
        if not file_urls:
            print(f"     ℹ️  No downloadable files found in HTML")
            # البحث في Wayback عن ملفات قديمة
            file_urls = await self._search_wayback_files(base)
        
        print(f"     📄 Found {len(file_urls)} files to analyze")
        
        # 3. تحليل كل ملف
        for file_url in file_urls[:20]:  # حد أقصى 20 ملف
            try:
                metadata = await self._analyze_file(client, file_url)
                if metadata and metadata.findings:
                    result.metadata_found.append(metadata)
                    result.total_usernames += len(metadata.usernames)
                    result.total_emails += len(metadata.email_addresses)
                    result.total_paths += len(metadata.internal_paths)
                    result.total_versions += len(metadata.software_versions)
            except Exception as e:
                logger.debug(f"Metadata analysis failed for {file_url}: {e}")
        
        result.files_analyzed = len(result.metadata_found)
        
        # 4. تجميع النتائج
        all_usernames = set()
        all_emails = set()
        all_paths = set()
        all_versions = set()
        
        for meta in result.metadata_found:
            all_usernames.update(meta.usernames)
            all_emails.update(meta.email_addresses)
            all_paths.update(meta.internal_paths)
            all_versions.update(meta.software_versions)
        
        # عرض النتائج
        if result.metadata_found:
            print(f"     ✅ Analyzed {result.files_analyzed} files")
            if all_usernames:
                print(f"     👤 Usernames: {', '.join(list(all_usernames)[:5])}")
            if all_emails:
                print(f"     📧 Emails: {len(all_emails)}")
            if all_paths:
                print(f"     📁 Internal paths: {len(all_paths)}")
            if all_versions:
                print(f"     📦 Software versions: {', '.join(list(all_versions)[:5])}")
        else:
            print(f"     ℹ️  No metadata found in files")
        
        self._results[base] = result
        return result
    
    def _extract_file_urls(self, html: str, base: str) -> List[str]:
        """استخراج روابط الملفات من HTML"""
        urls = set()
        
        # روابط مباشرة
        link_pattern = re.compile(r'(?:href|src)=["\']([^"\']+)["\']', re.I)
        for match in link_pattern.finditer(html):
            href = match.group(1)
            
            # فحص نوع الملف
            for pattern in self.FILE_PATTERNS:
                if re.search(pattern, href, re.I):
                    full_url = urljoin(base, href)
                    urls.add(full_url)
                    break
        
        return list(urls)[:30]
    
    async def _search_wayback_files(self, base: str) -> List[str]:
        """البحث عن ملفات قديمة في Wayback Machine"""
        urls = []
        
        try:
            async with httpx.AsyncClient(timeout=30, verify=False) as client:
                cdx_url = "https://web.archive.org/cdx/search/cdx"
                params = {
                    "url": f"{base}/*",
                    "output": "json",
                    "fl": "original",
                    "filter": "mimetype:application/pdf",
                    "limit": "50",
                }
                
                response = await client.get(cdx_url, params=params)
                if response.status_code == 200:
                    data = response.json()
                    for row in data[1:]:  # تجاهل header
                        if row and len(row) > 0:
                            urls.append(row[0])
        except:
            pass
        
        return urls[:20]
    
    async def _analyze_file(self, client: httpx.AsyncClient, file_url: str) -> Optional[MetadataInfo]:
        """تحليل ملف واحد"""
        try:
            # HEAD request للتحقق من الحجم
            head_response = await client.head(file_url)
            content_type = head_response.headers.get("content-type", "")
            content_length = int(head_response.headers.get("content-length", 0))
            
            # تجاهل الملفات الكبيرة جداً (> 50MB)
            if content_length > 50 * 1024 * 1024:
                return None
            
            # تجاهل الملفات الصغيرة جداً (< 100 bytes)
            if content_length < 100:
                return None
            
            # جلب الملف
            response = await client.get(file_url)
            if response.status_code != 200:
                return None
            
            content = response.content
            text_content = ""
            
            # استخراج النص حسب نوع الملف
            file_type = self._get_file_type(file_url, content_type)
            
            if file_type == "pdf":
                text_content = self._extract_pdf_text(content)
            elif file_type in ["docx", "xlsx", "pptx"]:
                text_content = self._extract_office_text(content, file_type)
            elif file_type in ["image"]:
                text_content = self._extract_image_exif(content)
            else:
                # نص عادي
                try:
                    text_content = content.decode('utf-8', errors='ignore')
                except:
                    text_content = content.decode('latin-1', errors='ignore')
            
            if not text_content:
                return None
            
            metadata = MetadataInfo(
                file_url=file_url,
                file_type=file_type,
                file_size=content_length,
                raw_text=text_content[:5000],
            )
            
            # استخراج البيانات
            self._extract_metadata_from_text(metadata, text_content)
            
            return metadata
            
        except Exception as e:
            logger.debug(f"File analysis failed for {file_url}: {e}")
            return None
    
    def _get_file_type(self, url: str, content_type: str) -> str:
        """تحديد نوع الملف"""
        url_lower = url.lower()
        ct_lower = content_type.lower()
        
        if '.pdf' in url_lower or 'pdf' in ct_lower:
            return "pdf"
        if '.doc' in url_lower or 'word' in ct_lower or 'document' in ct_lower:
            return "docx"
        if '.xls' in url_lower or 'excel' in ct_lower or 'spreadsheet' in ct_lower:
            return "xlsx"
        if '.ppt' in url_lower or 'powerpoint' in ct_lower or 'presentation' in ct_lower:
            return "pptx"
        if re.search(r'\.(?:jpg|jpeg|png|gif|tiff?|bmp)$', url_lower) or 'image' in ct_lower:
            return "image"
        return "other"
    
    def _extract_pdf_text(self, content: bytes) -> str:
        """استخراج النص من PDF"""
        text = ""
        
        try:
            # محاولة استخدام PyPDF2
            from PyPDF2 import PdfReader
            
            pdf_file = io.BytesIO(content)
            reader = PdfReader(pdf_file)
            
            for page in reader.pages[:5]:  # أول 5 صفحات
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            
            # استخراج metadata من PDF
            if reader.metadata:
                text += f"\nPDF Metadata:\n"
                for key, value in reader.metadata.items():
                    if value:
                        text += f"{key}: {value}\n"
        
        except ImportError:
            # Fallback: استخراج نص بسيط
            try:
                text = content.decode('utf-8', errors='ignore')
                # استخراج النص بين stream/endstream
                stream_pattern = re.compile(rb'stream\r?\n(.*?)\r?\nendstream', re.DOTALL)
                for match in stream_pattern.finditer(content):
                    try:
                        text += match.group(1).decode('utf-8', errors='ignore') + "\n"
                    except:
                        pass
            except:
                pass
        
        return text
    
    def _extract_office_text(self, content: bytes, file_type: str) -> str:
        """استخراج النص من مستندات Office"""
        text = ""
        
        try:
            # محاولة استخدام python-docx / openpyxl
            if file_type == "docx":
                from docx import Document
                doc = Document(io.BytesIO(content))
                for para in doc.paragraphs[:50]:
                    text += para.text + "\n"
                
                # Core properties
                if doc.core_properties:
                    cp = doc.core_properties
                    text += f"\nAuthor: {cp.author}\n"
                    text += f"Last Modified By: {cp.last_modified_by}\n"
                    text += f"Created: {cp.created}\n"
                    text += f"Modified: {cp.modified}\n"
            
            elif file_type == "xlsx":
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content), read_only=True)
                ws = wb.active
                for row in ws.iter_rows(max_row=20, values_only=True):
                    text += ' '.join([str(c) for c in row if c]) + "\n"
        
        except ImportError:
            # Fallback: استخراج XML من ZIP
            import zipfile
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as zf:
                    # docProps/core.xml
                    if 'docProps/core.xml' in zf.namelist():
                        core_xml = zf.read('docProps/core.xml').decode('utf-8', errors='ignore')
                        text += core_xml + "\n"
                    
                    # word/document.xml
                    if 'word/document.xml' in zf.namelist():
                        doc_xml = zf.read('word/document.xml').decode('utf-8', errors='ignore')
                        # استخراج النص بين <w:t> tags
                        text_parts = re.findall(r'<w:t[^>]*>(.*?)</w:t>', doc_xml)
                        text += ' '.join(text_parts[:200])
            except:
                pass
        
        return text
    
    def _extract_image_exif(self, content: bytes) -> str:
        """استخراج EXIF من الصور"""
        text = ""
        
        try:
            from PIL import Image
            from PIL.ExifTags import TAGS, GPSTAGS
            
            img = Image.open(io.BytesIO(content))
            exif_data = img._getexif()
            
            if exif_data:
                for tag_id, value in exif_data.items():
                    tag_name = TAGS.get(tag_id, tag_id)
                    if tag_name in ['Make', 'Model', 'Software', 'DateTime',
                                   'Artist', 'Copyright', 'ImageDescription',
                                   'GPSInfo']:
                        text += f"{tag_name}: {value}\n"
                    
                    if tag_name == 'GPSInfo' and isinstance(value, dict):
                        for gps_tag, gps_value in value.items():
                            gps_name = GPSTAGS.get(gps_tag, gps_tag)
                            text += f"GPS {gps_name}: {gps_value}\n"
        except ImportError:
            pass
        except Exception:
            pass
        
        return text
    
    def _extract_metadata_from_text(self, metadata: MetadataInfo, text: str):
        """استخراج البيانات من النص"""
        
        # أسماء المستخدمين
        for pattern in self.USERNAME_PATTERNS:
            matches = re.findall(pattern, text, re.I)
            for match in matches:
                if match and len(match) > 2 and len(match) < 50:
                    if match not in metadata.usernames:
                        metadata.usernames.append(match)
        
        # المسارات الداخلية
        for pattern in self.PATH_PATTERNS:
            matches = re.findall(pattern, text, re.I)
            for match in matches:
                if match not in metadata.internal_paths:
                    metadata.internal_paths.append(match)
        
        # إصدارات البرمجيات
        for pattern in self.VERSION_PATTERNS:
            matches = re.findall(pattern, text, re.I)
            for match in matches:
                if match not in metadata.software_versions:
                    metadata.software_versions.append(match)
        
        # إيميلات
        email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        emails = re.findall(email_pattern, text)
        metadata.email_addresses = list(set(emails))[:20]
        
        # URLs
        url_pattern = r'https?://[^\s<>"\']+'
        urls = re.findall(url_pattern, text)
        metadata.urls = list(set(urls))[:20]
        
        # تجميع النتائج
        if metadata.usernames:
            metadata.findings.append(f"Found {len(metadata.usernames)} usernames: {', '.join(metadata.usernames[:5])}")
        if metadata.email_addresses:
            metadata.findings.append(f"Found {len(metadata.email_addresses)} email addresses")
        if metadata.internal_paths:
            metadata.findings.append(f"Found {len(metadata.internal_paths)} internal paths")
        if metadata.software_versions:
            metadata.findings.append(f"Found {len(metadata.software_versions)} software versions")
    
    def get_results(self, url: str) -> Optional[MetadataResult]:
        return self._results.get(url)
    
    async def close(self):
        if self._client:
            await self._client.aclose()


# نسخة عالمية
_metadata_analyzer = None

def get_metadata_analyzer() -> MetadataAnalyzer:
    global _metadata_analyzer
    if _metadata_analyzer is None:
        _metadata_analyzer = MetadataAnalyzer()
    return _metadata_analyzer
